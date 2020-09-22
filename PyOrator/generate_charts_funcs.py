#-------------------------------------------------------------------------------
# Name:        generate_charts_funcs.py.py
# Purpose:
# Author:      Mike Martin
# Created:     07/03/2015
# Licence:     <your licence>
#-------------------------------------------------------------------------------
#!/usr/bin/env python

__prog__ = 'generate_charts_funcs.py'
__version__ = '0.0.1'

# Version history
# ---------------
# 0.0.1  Wrote.
#
import os
from string import ascii_uppercase
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)

preferred_line_width = 25000       # 100020 taken from chart_example.py     width in EMUs

def generate_charts(fname, data_frame, sheet_name):
    '''
    add charts to an existing Excel file
    '''
    func_name =  __prog__ + ' generate_charts'

    #  name of comparisons
    # ==============================
    min_num_cols = 26
    if not os.path.exists(fname):
        print('File ' + fname + ' must exist')
        return -1

    # wrkbk = Workbook()
    wb_obj = load_workbook(fname, data_only=True)
    sheet = wb_obj[sheet_name]
    if sheet.max_column < min_num_cols:
        print('Sheet ' + sheet_name + ' must must have at least {} columns, found: {}'
                                                    .format(min_num_cols, sheet.max_column))
        return -1

    alphabet_string = ascii_uppercase
    alphabet = list(alphabet_string)

    chart_sheet = wb_obj.create_sheet('charts')
    '''
    for row in range(2, sheet.max_row + 1):
        tstep_cell = 'D' + str(row)
        total_cell = 'Y' + str(row)
        chart_sheet.append([sheet[total_cell].value])
    '''
    nrow_chart = 10

    # generate charts for all metrics except for period, month and tstep
    # ==================================================================
    for col_indx in range(sheet.max_column, 4, -1):
        metric_chart = LineChart()
        metric_chart.style = 13

        metric = sheet[alphabet[col_indx - 1] + '1'].value      # TODO: find better way
        metric_chart.title = metric
        metric_chart.y_axis.title = 'kgC/ha'
        metric_chart.y_axis.title = 'tonnes C/ha'
        metric_chart.x_axis.title = 'Time step'
        metric_chart.height = 10
        metric_chart.width = 20

        data = Reference(sheet, min_col = col_indx, min_row = 1, max_col = col_indx, max_row = sheet.max_row)
        metric_chart.add_data(data, titles_from_data = True)

        # Style the lines
        # ===============
        sref = metric_chart.series[0]
        sref.graphicalProperties.line.width = preferred_line_width
        sref.graphicalProperties.line.solidFill = "FF0000"
        sref.smooth = True

        # now write to previously created sheet
        # =====================================
        chart_sheet.add_chart(metric_chart, "D" + str(nrow_chart))
        nrow_chart += 20

    try:
        wb_obj.active = 1   # make the charts sheet active - assumes there are only two sheets
        wb_obj.save(fname)
        print('\tcreated: ' + fname)
    except PermissionError as e:
        print(str(e) + ' - could not create: ' + fname)

    return
