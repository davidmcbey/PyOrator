# -------------------------------------------------------------------------------
# Name:        ora_excel_read.py
# Purpose:     a collection of reusable functions
# Author:      Mike Martin
# Created:     26/12/2019
# Licence:     <your licence>
# Definitions:
#   spin_up
#
# Description:
#
#
# -------------------------------------------------------------------------------
# !/usr/bin/env python

__prog__ = 'ora_excel_read.py'
__version__ = '0.0.0'

# Version history
# ---------------
#
import os
from openpyxl import load_workbook
from pandas import Series, read_excel, DataFrame
from zipfile import BadZipFile
from ora_excel_write import check_out_dir, retrieve_output_xls_files
from ora_water_model import add_pet_to_weather
from ora_low_level_fns import average_weather

METRIC_LIST = list(['precip', 'tair'])
MNTH_NAMES_SHORT = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
REQUIRED_SHEET_NAMES = list(['Inputs1- Farm location', 'N constants', 'Inputs3- Soils & Crops',
                             'Inputs3b- Soils & Rotations', 'Inputs3d- Changes in rotations', 'Crop parms',
                             'Org Waste parms', 'Weather', 'Inputs4- Livestock','C1. Change in animal production',
                             'C1a. Typical animal production'])

ERR_MESS_SHEET = '*** Error *** reading sheet '


def check_excel_input_file(form, xls_inp_fname):
    '''
    validate selected Excel file
    '''
    fileOkFlag = True

    if not os.path.isfile(xls_inp_fname):
        return None

    form.w_study.setText('Study not set')
    form.settings['inp_dir'] = ''
    form.w_soil_cn.setEnabled(False)
    form.w_disp_out.setEnabled(False)

    print('Loading: ' + xls_inp_fname)
    try:
        wb_obj = load_workbook(xls_inp_fname, data_only=True)
        sheet_names = wb_obj.sheetnames
    except (PermissionError, BadZipFile) as e:
        print('Error: ' + str(e))
        return None

    wb_obj.close()

    # all required sheets must be present
    # ===================================
    for sheet in REQUIRED_SHEET_NAMES:
        if sheet not in sheet_names:
            fileOkFlag = False
            break

    if fileOkFlag:
        mess = 'Excel input file is valid'
        form.w_soil_cn.setEnabled(True)

        study, latitude, longitude = _read_location_sheet(xls_inp_fname, 'Inputs1- Farm location', 13)
        study_desc = 'Study: ' + study
        study_desc += '\tLatitude: {}'.format(latitude)
        study_desc += '\tLongitude: {}'.format(longitude)
        form.w_study.setText(study_desc)

        form.settings['study'] = study
        form.settings['inp_dir'], dummy = os.path.split(xls_inp_fname)
        check_out_dir(form)
        retrieve_output_xls_files(form, study)

    else:
        mess = 'Required sheet ' + sheet + ' is not present - please check file'

    print(mess + '\n')
    return mess


def _read_location_sheet(xls_fname, sheet_name, skip_until):
    nlines_crop_desc = 26  # temporary
    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until), usecols=range(2, 4))

    location = data.dropna(how='all')
    location = location.rename(columns={'Unnamed: 2': 'Description', 'Unnamed: 3': 'Value'})

    study = location['Value'].values[1]
    latitude = location['Value'].values[2]
    longitude = location['Value'].values[3]

    return study, latitude, longitude


def _read_n_constants_sheet(xls_fname, sheet_name, skip_until):
    '''
    r_dry is an environmental constant
    '''

    n_parm_names = list(['atmos_n_depos', 'prop_atmos_dep_no3', 'no3_min', 'k_nitrif',
                         'n_denit_max', 'n_d50', 'prop_n2o_fc', 'prop_nitrif_gas', 'prop_nitrif_no',
                         'precip_critic', 'prop_volat', 'prop_atmos_dep_nh4', 'c_n_rat_som', 'r_dry'])

    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until), usecols=range(1, 3))
    n_parms_df = data.dropna(how='all')
    n_parms = {}
    for indx, defn in enumerate(n_parm_names):
        n_parms[defn] = n_parms_df['Value'].values[indx]

    return n_parms


def _read_crop_mngmnt_sheet(xls_fname, sheet_name, skip_until, crop_vars):
    nlines_crop_desc = 26  # temporary
    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until), usecols=range(1, 9))
    management = data.dropna(how='all')
    area_names = management.columns[3:]
    management = management.rename(columns={'Unnamed: 1': 'Crop'})

    soil_section = management['Crop'].loc[management['Crop'] == 'SOILS']

    crop_starts = management['Crop'].loc[management['Crop'] == 'Crop ']

    crop_mngmnt_all_areas = {}
    soil_all_areas = {}

    for area in area_names:

        if len(soil_section) > 0:
            soil_all_areas[area] = Soil(management[area].values[2: 10])

        crop_defns = []
        for indx in crop_starts.index:
            crop_slice = management[area].values[indx + 1: indx + nlines_crop_desc - 1]
            crop_defn = Crop(crop_slice)
            crop_name = crop_defn.crop_lu
            if type(crop_name) is not str:
                break

            # check for consistency
            # =====================
            crop_defns.append(crop_defn)
            if crop_name in crop_vars:
                ngrow = crop_defn.harvest_mnth - crop_defn.sowing_mnth + 1
                nprops = len(crop_vars[crop_name]['pi_prop'])
                if ngrow != nprops:
                    print(ERR_MESS_SHEET + sheet_name + ' inconsistent growing season lengths: {} {}'.format(ngrow,
                                                                                                             nprops))
            else:
                print(ERR_MESS_SHEET + sheet_name + ' unknown crop: {} '.format(crop_name))

        crop_mngmnt_all_areas[area] = crop_defns

    return crop_mngmnt_all_areas, soil_all_areas


def _extract_weather(pettmp_dframe, year_strt, year_end, indx):
    pettemp = []
    for yr in range(year_strt, year_end):
        slice = list(pettmp_dframe[yr][indx:indx + 12])
        pettemp += slice

    return pettemp


def _read_weather_sheet(xls_fname, sheet_name, skip_until):
    '''
    reads weather
    '''
    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until), usecols=range(3, 16))
    pettmp_dframe = data.dropna(how='all')
    year_strt = pettmp_dframe.columns[2]
    ncols = len(pettmp_dframe.columns)
    nyears = ncols - 3
    year_end = year_strt + nyears
    col_names = ['Month', 'Season'] + [year_strt + val for val in range(nyears)] + ['dummy']
    pettmp_dframe.columns = col_names

    # reorganise into total rainfall (mm month-1) and average air temperature (°C) for steady state and forward runs
    # ==============================================================================================================
    pettmp_ss = {}
    for indx, metric in zip(list([0, 14]), METRIC_LIST):
        pettmp_ss[metric] = _extract_weather(pettmp_dframe, year_strt, year_end, indx)

    pettmp_fwd = {}
    for indx, metric in zip(list([28, 42]), METRIC_LIST):
        pettmp_fwd[metric] = _extract_weather(pettmp_dframe, year_strt, year_end, indx)

    return pettmp_ss, pettmp_fwd


def _read_crop_vars(xls_fname, sheet_name):
    '''
    read maximum rooting depths etc. for each crop from sheet A1c
    '''
    crop_parm_names = Series(list(['lu_code', 'rat_dpm_rpm', 'harv_indx', 'prop_npp_2soil', 'max_root_dpth',
                                   'max_root_dpth_orig', 'sow_mnth', 'harv_mnth', 'dummy1']) + MNTH_NAMES_SHORT + list(
        ['dummy2',
         'c_n_rat_pi', 'n_supply_min', 'n_supply_opt', 'n_respns_coef', 'fert_use_eff']))

    data = read_excel(xls_fname, sheet_name)
    data = data.dropna(how='all')
    try:
        crop_dframe = data.set_index(crop_parm_names)
        crop_vars = crop_dframe.to_dict()
    except ValueError as err:
        print(ERR_MESS_SHEET + sheet_name + ' ' + err)
        crop_vars = None

    # normalise PI tonnages
    # =====================
    for crop in ['Crop', 'None', 'Null']:
        del (crop_vars[crop])

    for crop_name in crop_vars:
        pi_tonnes = []
        for mnth in MNTH_NAMES_SHORT:
            pi = crop_vars[crop_name][mnth]
            if pi > 0:
                pi_tonnes.append(pi)

        pi_proportions = [val / sum(pi_tonnes) for val in pi_tonnes]
        crop_vars[crop_name]['pi_tonnes'] = pi_tonnes
        crop_vars[crop_name]['pi_prop'] = pi_proportions

        crop_vars[crop_name]['nmnths_grow'] = crop_vars[crop_name]['harv_mnth'] - crop_vars[crop_name]['sow_mnth'] + 1

    return crop_vars


def _read_organic_waste_sheet(xls_fname, sheet_name, skip_until):
    '''
    read Organic waste parameters
    added  - see (eq.2.1.12) and (eq.2.1.13)
    TODO percentages are converted to fraction
    '''
    ow_parms_names = Series(list(['c_n_rat', 'prop_nh4', 'rat_dpm_hum_ow', 'prop_iom_ow', 'pcnt_c', 'min_e_pcnt_wd',
                                  'max_e_pcnt_wd', 'ann_c_input', 'pcnt_urea']))

    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until))
    data = data.dropna(how='all')
    try:
        ow_dframe = data.set_index(ow_parms_names)
        all_ow_parms = ow_dframe.to_dict()
    except ValueError as err:
        print(ERR_MESS_SHEET + sheet_name + ' ' + str(err))
        all_ow_parms = None

    return all_ow_parms


def read_livestock_data(xls_fname):
    """A function to pull livestock data from excel file, and store as
    pandas dataframe"""
    excel_file = xls_fname
    data = read_excel(excel_file, sheet_name="Inputs4- Livestock",
                      usecols=range(3, 10), skiprows=range(0, 17))
    livestock_input = DataFrame(data)

    return livestock_input


def read_livestock_supp_data(xls_fname):
    """A function to pull the data on location etc. from sheet C1 of Orator, and
    store as a dictionary"""
    wb = load_workbook(xls_fname, data_only=True)
    input_supp_info = wb["C1. Change in animal production"]
    farm_info = {'Region': input_supp_info['C18'].value,
                 'Production': input_supp_info['F18'].value,
                 'Climate': input_supp_info['H18'].value,
                 'System': input_supp_info['K18'].value
                 }

    return farm_info


def read_africa_animal_prod(xls_fname):
    """A function to pull African animal production data from excel file,
    and store as pandas dataframe"""
    global africa_animal_prod
    excel_file = xls_fname
    data = read_excel(excel_file, sheet_name="C1a. Typical animal production",
                      usecols=range(1, 15), skiprows=range(0, 12))
    africa_animal_prod = DataFrame(data)

    return africa_animal_prod

def read_single_crop_and_soil_data(xls_fname):
    """Pull data from ORATOR Excel sheet 'Inputs3- Soils & Crops'"""
    excel_file = xls_fname
    data = read_excel(excel_file, sheet_name="Inputs3- Soils & Crops",
                                usecols= 'B, D:J, L, M, N', skiprows=range(0,13))
    single_crop_data = DataFrame(data)

    return single_crop_data

def read_rotations_crop_and_soil_data(xls_fname):
    """Pull data from ORATOR Excel Sheet 'Inputs3b- Soils & Rotations'"""
    excel_file = xls_fname
    data = read_excel(excel_file, sheet_name='Inputs3b- Soils & Rotations',
                                usecols= 'E:I', skiprows=range(0,12))
    crop_rotation_data = DataFrame(data)

    return crop_rotation_data


def read_weather_data_crops(xls_fname):
    """Get steady state and forward run air temp and rainfall data from ORATOR excel sheet 'Inputs2- Weather'
    This is duplication of _read_weather_sheet above, can be unified later """
    excel_file = xls_fname
    data = read_excel(excel_file, sheet_name="Weather",
                         usecols='F:O', skiprows=range(0, 14),)
    weather_data = DataFrame(data)
    weather_data.columns = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
    return weather_data


class ReadStudy(object, ):

    def __init__(self, xls_inp_fname, out_dir, output_excel=True):
        '''
        read location sheet from ORATOR inputs Excel file
        '''

        self.output_excel = output_excel
        self.out_dir = out_dir

        # Farm location
        # =============
        self.study_name, self.latitude, self.longitude \
            = _read_location_sheet(xls_inp_fname, 'Inputs1- Farm location', 13)


class ReadWeather(object, ):

    def __init__(self, xls_inp_fname, latitude):
        '''
        read parameters from ORATOR inputs Excel file
        '''

        print('Reading weather sheet...')

        pettmp_ss, pettmp_fwd = _read_weather_sheet(xls_inp_fname, 'Weather', 14)

        # generate PET from weather
        # =========================
        self.pettmp_ss = add_pet_to_weather(latitude, pettmp_ss)
        self.pettmp_fwd = add_pet_to_weather(latitude, pettmp_fwd)

        # average monthly precip and temp is required for spin up
        # =======================================================
        self.ave_precip_ss, self.ave_temp_ss, self.ave_pet_ss = \
            average_weather(latitude, self.pettmp_ss['precip'], self.pettmp_ss['tair'])

        # get average annual rain and temperature of first 10 years
        # =========================================================
        nmnths = len(pettmp_ss['precip'])
        nyrs = nmnths / 12
        self.ann_ave_precip_ss = sum(pettmp_ss['precip']) / nyrs
        self.ann_ave_temp_ss = sum(pettmp_ss['tair']) / nmnths


class ReadInputParms(object, ):

    def __init__(self, xls_inp_fname):
        '''
        read parameters from ORATOR inputs Excel file
        '''

        print('Reading parameter sheets...')

        # Nitrogen params plus r_dry, drying potential
        # ============================================
        self.n_parms = _read_n_constants_sheet(xls_inp_fname, 'N constants', 0)

        # Organic Waste and Crop params e.g. max rooting depths
        # =====================================================
        self.ow_parms = _read_organic_waste_sheet(xls_inp_fname, 'Org Waste parms', 0)
        self.crop_vars = _read_crop_vars(xls_inp_fname, 'Crop parms')


class ReadInputSubplots(object, ):

    def __init__(self, xls_inp_fname, crop_vars):
        '''
        read parameters from ORATOR inputs Excel file
        '''

        print('Reading management sheets...')

        # Soil params and management
        # ==========================
        self.crop_mngmnt_ss, self.soil_all_areas = \
            _read_crop_mngmnt_sheet(xls_inp_fname, 'Inputs3b- Soils & Rotations', 13, crop_vars)
        self.crop_mngmnt_fwd, dummy = \
            _read_crop_mngmnt_sheet(xls_inp_fname, 'Inputs3d- Changes in rotations', 14, crop_vars)


class Soil(object, ):
    '''

    '''

    def __init__(self, soil_slice):
        """
        Assumptions:
        """
        self.title = 'Soil'

        t_depth = soil_slice[0]
        self.t_clay = soil_slice[1]
        self.t_silt = soil_slice[2]
        self.t_sand = soil_slice[3]
        t_carbon = soil_slice[4]
        t_bulk = soil_slice[5]
        self.t_pH_h2o = soil_slice[6]
        self.t_salinity = soil_slice[7]

        tot_soc_meas = (10 ** 4) * (t_depth / 100) * t_bulk * (t_carbon / 100)  # tonnes/ha

        self.t_depth = t_depth
        self.t_carbon = t_carbon
        self.t_bulk = t_bulk
        self.tot_soc_meas = tot_soc_meas


class Crop(object, ):
    '''

    '''

    def __init__(self, crop_slice):
        """
        Assumptions:
        """
        self.title = 'Crop'

        self.crop_lu = crop_slice[0]
        self.sowing_mnth = crop_slice[1]
        self.harvest_mnth = crop_slice[2]
        self.typical_yield = crop_slice[3]

        # inorganic fertiliser application
        # ================================
        self.fert_type = crop_slice[5]
        self.fert_n = crop_slice[6]
        self.fert_p = crop_slice[7]
        self.fert_mnth = crop_slice[8]

        # organic waste
        # =============
        self.ow_type = crop_slice[10]
        self.ow_mnth = crop_slice[11]
        self.ow_amount = crop_slice[12]

        # irrigation up to 12 months max
        # ==============================
        indx_strt = 14
        irrig = {}
        nmnths = int(len(crop_slice[indx_strt:]) / 2)
        for imnth in range(nmnths):
            indx = indx_strt + imnth * 2
            mnth = crop_slice[indx]
            amount = crop_slice[indx + 1]
            if amount == 0 or mnth == 0:
                continue
            else:
                irrig[mnth] = amount

        self.irrig = irrig
